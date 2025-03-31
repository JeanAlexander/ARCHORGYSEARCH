import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox, scrolledtext
from openpyxl import load_workbook
import re

def seleccionar_archivo():
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    if archivo:
        entry_ruta.delete(0, tk.END)
        entry_ruta.insert(0, archivo)

def contar_palabras_excel(ruta_excel):
    df = pd.read_excel(ruta_excel, sheet_name=None)
    total_palabras = 0
    inventario = {}

    for nombre_hoja, datos in df.items():
        palabras_hoja = sum(datos.astype(str).apply(lambda x: x.str.count(r"\b\w+\b")).sum())
        total_palabras += palabras_hoja
        inventario[nombre_hoja] = palabras_hoja

    return total_palabras, inventario

def buscar_palabra_en_hojas(ruta_excel):
    palabra = simpledialog.askstring("Buscar palabra", "Ingresa la palabra a buscar:")
    if not palabra:
        return
    
    df = pd.read_excel(ruta_excel, sheet_name=None)
    hojas_con_palabra = [nombre_hoja for nombre_hoja, datos in df.items()
                          if datos.astype(str).apply(lambda x: x.str.contains(palabra, case=False, na=False)).any().any()]
    
    if hojas_con_palabra:
        mensaje = f"🔍 La palabra '{palabra}' fue encontrada en:\n" + "\n".join(f"📄 {hoja}" for hoja in hojas_con_palabra)
    else:
        mensaje = f"La palabra '{palabra}' no se encontró en ninguna hoja."
    
    mostrar_texto_en_ventana("Resultado de Búsqueda", mensaje)

def organizar_hojas_por_anio():
    archivo = entry_ruta.get()
    if not archivo:
        messagebox.showerror("Error", "Seleccione un archivo Excel primero.")
        return
    
    try:
        libro = load_workbook(archivo)
        hojas_ordenadas = {}

        for hoja in libro.sheetnames:
            ws = libro[hoja]
            valor_h9 = ws["H9"].value
            año = int(re.search(r"\d{4}", str(valor_h9)).group()) if valor_h9 and re.search(r"\d{4}", str(valor_h9)) else 9999
            hojas_ordenadas[hoja] = año
        
        hojas_ordenadas = dict(sorted(hojas_ordenadas.items(), key=lambda item: item[1]))
        libro._sheets.sort(key=lambda sheet: hojas_ordenadas.get(sheet.title, 9999))
        archivo_guardado = archivo.replace(".xlsx", "_ordenado.xlsx")
        libro.save(archivo_guardado)
        messagebox.showinfo("Éxito", f"Hojas organizadas y guardadas en:\n{archivo_guardado}")
    
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un problema: {e}")

def mostrar_conteo_palabras():
    archivo = entry_ruta.get()
    if not archivo:
        messagebox.showerror("Error", "Seleccione un archivo Excel primero.")
        return
    
    total, inventario = contar_palabras_excel(archivo)
    mensaje = f"📊 Total de palabras en el archivo: {total}\n\n" + "\n".join(f"📄 {hoja}: {cantidad} palabras" for hoja, cantidad in inventario.items())
    mostrar_texto_en_ventana("Conteo de Palabras", mensaje)

def mostrar_texto_en_ventana(titulo, contenido):
    ventana = tk.Toplevel(root)
    ventana.title(titulo)
    ventana.geometry("500x400")
    text_area = scrolledtext.ScrolledText(ventana, wrap=tk.WORD, width=60, height=20)
    text_area.insert(tk.INSERT, contenido)
    text_area.config(state=tk.DISABLED)
    text_area.pack(pady=10, padx=10)

# Interfaz gráfica
root = tk.Tk()
root.title("Gestor de Archivos Excel")
root.geometry("520x400")
root.configure(bg="#f8f9fa")

frame = tk.Frame(root, bg="#f8f9fa")
frame.pack(pady=20)

tk.Label(frame, text="📂 Seleccionar archivo Excel:", bg="#f8f9fa", font=("Arial", 12, "bold")).pack(pady=5)
entry_ruta = tk.Entry(frame, width=50)
entry_ruta.pack(pady=5)
tk.Button(frame, text="🔍 Buscar", command=seleccionar_archivo, bg="#007bff", fg="white", font=("Arial", 10)).pack(pady=5)

tk.Button(frame, text="📊 Contar Palabras", command=mostrar_conteo_palabras, bg="#28a745", fg="white", font=("Arial", 10)).pack(pady=5)
tk.Button(frame, text="🔎 Buscar Palabra", command=lambda: buscar_palabra_en_hojas(entry_ruta.get()), bg="#ffc107", font=("Arial", 10)).pack(pady=5)
tk.Button(frame, text="📑 Organizar Hojas por Año", command=organizar_hojas_por_anio, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

tk.Label(root, text="🔹 La organización de hojas por año usa el valor de la celda H9 en cada hoja.\nSi la celda contiene un año (Ejemplo: 2023), se ordenará según ese valor.",
         bg="#f8f9fa", fg="#6c757d", font=("Arial", 10), wraplength=500, justify="center").pack(pady=10)

tk.Button(root, text="🚪 Salir", command=root.quit, bg="#6c757d", fg="white", font=("Arial", 10)).pack(pady=10)

root.mainloop()
